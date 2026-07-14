import openpyxl
import xlrd
import csv
import os
import glob

base = os.path.dirname(os.path.abspath(__file__))
out_dir = os.path.join(base, 'csv_data')
os.makedirs(out_dir, exist_ok=True)

# 1. Convert root-level xlsx files
root_files = glob.glob(os.path.join(base, '*.xlsx')) + glob.glob(os.path.join(base, '*.xls'))
for fp in root_files:
    fname = os.path.basename(fp)
    print(f'Converting: {fname}')
    try:
        if fname.endswith('.xls'):
            wb = xlrd.open_workbook(fp)
            for sname in wb.sheet_names():
                ws = wb.sheet_by_name(sname)
                csv_name = f'{os.path.splitext(fname)[0]}_{sname}.csv'
                with open(os.path.join(out_dir, csv_name), 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    for r in range(ws.nrows):
                        writer.writerow([ws.cell_value(r, c) for c in range(ws.ncols)])
        else:
            wb = openpyxl.load_workbook(fp, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                csv_name = f'{os.path.splitext(fname)[0]}_{sname}.csv'
                with open(os.path.join(out_dir, csv_name), 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(row)
    except Exception as e:
        print(f'  ERROR: {e}')

# 2. Convert yearbook files from a few key years
for yr in range(2003, 2023):
    folder = None
    for d in os.listdir(base):
        if d.startswith(str(yr)) and os.path.isdir(os.path.join(base, d)):
            folder = os.path.join(base, d)
            break
    if not folder:
        continue
    yr_out = os.path.join(out_dir, str(yr))
    os.makedirs(yr_out, exist_ok=True)
    for fname in os.listdir(folder):
        fp = os.path.join(folder, fname)
        if not (fname.endswith('.xlsx') or fname.endswith('.xls')):
            continue
        print(f'Converting: {yr}/{fname}')
        try:
            if fname.endswith('.xls'):
                wb = xlrd.open_workbook(fp)
                ws = wb.sheet_by_index(0)
                csv_name = os.path.splitext(fname)[0] + '.csv'
                with open(os.path.join(yr_out, csv_name), 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    for r in range(ws.nrows):
                        writer.writerow([ws.cell_value(r, c) for c in range(ws.ncols)])
            else:
                wb = openpyxl.load_workbook(fp, data_only=True)
                ws = wb.active
                csv_name = os.path.splitext(fname)[0] + '.csv'
                with open(os.path.join(yr_out, csv_name), 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(row)
        except Exception as e:
            print(f'  ERROR: {e}')

# 3. Convert electricity 2020-2022
elec_dir = os.path.join(base, '用电量2020-2022')
if os.path.isdir(elec_dir):
    elec_out = os.path.join(out_dir, '用电量2020-2022')
    os.makedirs(elec_out, exist_ok=True)
    for fname in os.listdir(elec_dir):
        fp = os.path.join(elec_dir, fname)
        if fname.endswith('.png'):
            continue
        print(f'Converting: 用电量2020-2022/{fname}')
        try:
            if fname.endswith('.xls'):
                wb = xlrd.open_workbook(fp)
                ws = wb.sheet_by_index(0)
                csv_name = os.path.splitext(fname)[0] + '.csv'
                with open(os.path.join(elec_out, csv_name), 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    for r in range(ws.nrows):
                        writer.writerow([ws.cell_value(r, c) for c in range(ws.ncols)])
            else:
                wb = openpyxl.load_workbook(fp, data_only=True)
                ws = wb.active
                csv_name = os.path.splitext(fname)[0] + '.csv'
                with open(os.path.join(elec_out, csv_name), 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(row)
        except Exception as e:
            print(f'  ERROR: {e}')

print('\nDone! All CSV files saved to csv_data/')
